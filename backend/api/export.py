"""
Phase 12: Export endpoints — CSV, JSON, Excel.
All filters mirror the transactions list endpoint so the UI Export button
can pass the same active filter params through.
"""
import csv
import io
import json
from datetime import date
from decimal import Decimal
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_

from api.auth import get_current_user
from db.database import get_db
from db.models import Transaction, Account, Budget, Subscription

router = APIRouter(tags=["export"])

_HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ALT_FILL    = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")


# ── Shared helpers ─────────────────────────────────────────────────────────────

def _apply_filters(stmt, *, date_from, date_to, category, direction,
                   account_id, search, is_recurring, need_want_savings):
    conds = []
    if date_from:             conds.append(Transaction.date >= date_from)
    if date_to:               conds.append(Transaction.date <= date_to)
    if category:              conds.append(Transaction.category == category)
    if direction:             conds.append(Transaction.direction == direction)
    if account_id:            conds.append(Transaction.account_id.cast(str) == account_id)
    if is_recurring is not None:
                              conds.append(Transaction.is_recurring == is_recurring)
    if need_want_savings:     conds.append(Transaction.need_want_savings == need_want_savings)
    if search:
        q = f"%{search}%"
        conds.append(or_(
            Transaction.merchant.ilike(q),
            Transaction.description.ilike(q),
            Transaction.notes.ilike(q),
        ))
    if conds:
        stmt = stmt.where(and_(*conds))
    return stmt


def _f(v):
    """Float or empty string for nullable Decimal."""
    if v is None:
        return ""
    return float(v)


def _tx_row(t: Transaction, acct_map: dict) -> dict:
    return {
        "id":                     str(t.id),
        "date":                   str(t.date),
        "merchant":               t.merchant or t.description or "",
        "description":            t.description or "",
        "amount":                 _f(t.amount),
        "direction":              t.direction or "",
        "category":               t.category or "",
        "subcategory":            t.subcategory or "",
        "account":                acct_map.get(str(t.account_id), "") if t.account_id else "",
        "need_want":              t.need_want_savings or "",
        "fixed_var":              t.fixed_variable or "",
        "personal_work":          t.personal_work_shared or "",
        "notes":                  t.notes or "",
        "tags":                   ",".join(t.tags or []),
        "is_reimbursable":        t.is_reimbursable,
        "reimbursement_status":   t.reimbursement_status or "",
        "expected_reimbursement": _f(t.expected_reimbursement),
        "received_reimbursement": _f(t.received_reimbursement),
        "net_personal_cost":      _f(t.net_personal_cost),
        "is_recurring":           t.is_recurring,
        "source":                 t.source or "",
        "needs_review":           t.needs_review,
    }


CSV_FIELDS = [
    "id", "date", "merchant", "description", "amount", "direction",
    "category", "subcategory", "account", "need_want", "fixed_var",
    "personal_work", "notes", "tags", "is_reimbursable",
    "reimbursement_status", "expected_reimbursement", "received_reimbursement",
    "net_personal_cost", "is_recurring", "source", "needs_review",
]


async def _get_txs(db, **kw):
    stmt = _apply_filters(
        select(Transaction).order_by(Transaction.date.desc()), **kw
    )
    return (await db.execute(stmt)).scalars().all()


async def _get_acct_map(db) -> dict:
    return {str(a.id): a.name
            for a in (await db.execute(select(Account))).scalars().all()}


# ── CSV ────────────────────────────────────────────────────────────────────────

@router.get("/csv")
async def export_csv(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    category: Optional[str] = None,
    direction: Optional[str] = None,
    account_id: Optional[str] = None,
    search: Optional[str] = None,
    is_recurring: Optional[bool] = None,
    need_want_savings: Optional[str] = None,
    _user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export filtered transactions as CSV (matches the Transactions page filters)."""
    txs = await _get_txs(
        db, date_from=date_from, date_to=date_to, category=category,
        direction=direction, account_id=account_id, search=search,
        is_recurring=is_recurring, need_want_savings=need_want_savings,
    )
    acct_map = await _get_acct_map(db)

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=CSV_FIELDS)
    writer.writeheader()
    for t in txs:
        writer.writerow(_tx_row(t, acct_map))
    buf.seek(0)

    fname = f"transactions_{date.today().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        iter([buf.read()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── JSON ───────────────────────────────────────────────────────────────────────

@router.get("/json")
async def export_json(
    _user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export full dataset as JSON (transactions + accounts + budgets + subscriptions)."""
    txs      = await _get_txs(db, date_from=None, date_to=None, category=None,
                               direction=None, account_id=None, search=None,
                               is_recurring=None, need_want_savings=None)
    acct_map = await _get_acct_map(db)

    accounts = (await db.execute(select(Account))).scalars().all()
    budgets  = (await db.execute(
        select(Budget).order_by(Budget.year.desc(), Budget.month.desc())
    )).scalars().all()
    subs     = (await db.execute(select(Subscription))).scalars().all()

    def _d(v):
        return float(v) if isinstance(v, Decimal) else v

    payload = {
        "exported_at": date.today().isoformat(),
        "version": "1.0",
        "accounts": [
            {"id": str(a.id), "name": a.name, "type": a.type,
             "institution": a.institution, "last_four": a.last_four,
             "currency": a.currency, "is_active": a.is_active}
            for a in accounts
        ],
        "transactions": [
            {**_tx_row(t, acct_map), "id": str(t.id)} for t in txs
        ],
        "budgets": [
            {"month": b.month, "year": b.year, "category": b.category,
             "budget_amount": _d(b.budget_amount)}
            for b in budgets
        ],
        "subscriptions": [
            {"name": s.name, "amount": _d(s.amount),
             "billing_frequency": s.billing_frequency,
             "next_billing_date": str(s.next_billing_date) if s.next_billing_date else None,
             "category": s.category, "is_active": s.is_active,
             "value_rating": s.value_rating}
            for s in subs
        ],
    }

    fname = f"finance_export_{date.today().strftime('%Y%m%d')}.json"
    return StreamingResponse(
        iter([json.dumps(payload, indent=2)]),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Excel ──────────────────────────────────────────────────────────────────────

def _xl_header(ws, headers: list):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20


def _xl_autofit(ws, cap=50):
    for col in ws.columns:
        best = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(best + 3, cap)


@router.get("/excel")
async def export_excel(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    category: Optional[str] = None,
    direction: Optional[str] = None,
    account_id: Optional[str] = None,
    search: Optional[str] = None,
    _user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export transactions + budgets + subscriptions as a multi-sheet Excel workbook."""
    txs      = await _get_txs(db, date_from=date_from, date_to=date_to, category=category,
                               direction=direction, account_id=account_id, search=search,
                               is_recurring=None, need_want_savings=None)
    acct_map = await _get_acct_map(db)

    budgets = (await db.execute(
        select(Budget).order_by(Budget.year.desc(), Budget.month.desc())
    )).scalars().all()
    subs = (await db.execute(
        select(Subscription).where(Subscription.is_active == True)
    )).scalars().all()

    wb = openpyxl.Workbook()

    # ── Transactions sheet ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Transactions"
    _xl_header(ws, [
        "Date", "Merchant", "Amount", "Direction", "Category", "Subcategory",
        "Account", "Need/Want", "Fixed/Var", "Personal/Work",
        "Notes", "Reimbursable", "Net Cost", "Recurring", "Tags",
    ])
    ws.freeze_panes = "A2"

    for i, t in enumerate(txs, 2):
        fill = _ALT_FILL if i % 2 == 0 else None
        row_vals = [
            str(t.date),
            t.merchant or t.description or "",
            _f(t.amount),
            t.direction or "",
            t.category or "",
            t.subcategory or "",
            acct_map.get(str(t.account_id), "") if t.account_id else "",
            t.need_want_savings or "",
            t.fixed_variable or "",
            t.personal_work_shared or "",
            t.notes or "",
            "Yes" if t.is_reimbursable else "No",
            _f(t.net_personal_cost),
            "Yes" if t.is_recurring else "No",
            ", ".join(t.tags or []),
        ]
        for col, v in enumerate(row_vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            if fill:
                cell.fill = fill

    _xl_autofit(ws)

    # ── Budgets sheet ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Budgets")
    _xl_header(ws2, ["Year", "Month", "Category", "Budget Amount"])
    for i, b in enumerate(budgets, 2):
        for col, v in enumerate([b.year, b.month, b.category, float(b.budget_amount)], 1):
            ws2.cell(row=i, column=col, value=v)
    _xl_autofit(ws2)

    # ── Subscriptions sheet ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Subscriptions")
    _xl_header(ws3, [
        "Name", "Amount", "Frequency", "Next Billing",
        "Category", "Monthly Equiv", "Annual Equiv", "Value Rating",
    ])
    freq_months = {"monthly": 1, "yearly": 12, "quarterly": 3, "weekly": 0.25}
    for i, s in enumerate(subs, 2):
        m = freq_months.get(s.billing_frequency or "monthly", 1) or 1
        monthly = float(s.amount) / m
        for col, v in enumerate([
            s.name, float(s.amount), s.billing_frequency or "",
            str(s.next_billing_date) if s.next_billing_date else "",
            s.category or "", round(monthly, 2), round(monthly * 12, 2),
            s.value_rating or "",
        ], 1):
            ws3.cell(row=i, column=col, value=v)
    _xl_autofit(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"finance_export_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
